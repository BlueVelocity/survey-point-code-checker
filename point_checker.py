
from openpyxl import Workbook
from openpyxl import load_workbook
from easygui import fileopenbox, filesavebox, enterbox, ccbox, msgbox
from sys import exit
from datetime import *
import os
import re


def unidentified_error_handler(func):
    def wrapper(*args):
        try:
            return func(*args)
        except SystemExit:
            exit()
        except:
            msgbox(f'ExecutionError: See error log file located at: {os.getcwd()}/log.txt', 'Error Message', 'OK' )
            with open('./log.txt', 'w') as log:
                log.write(f'{datetime.now()}: ExecutionError: Action stopped at "{func.__name__}"\n')
            exit()
    return wrapper

@unidentified_error_handler
def prompt_continue(msg):
    title = 'Please confirm'
    if ccbox(msg, title, choices=('Continue', 'Exit')):
        pass
    else:
        exit()

@unidentified_error_handler
def prompt_string_input(msg):
    title = 'Please enter...'
    d_text = 'output'
    check_file_name_validity_pattern = r"^[A-Za-z0-9]+([-_][A-Za-z0-9]+)*(\.[A-Za-z0-9]+)?$"
    is_valid = False
    while(is_valid == False):
        string = enterbox(msg, title, d_text)
        if (re.match(check_file_name_validity_pattern, string) != None):
            is_valid = True
    return string        

@unidentified_error_handler
def select_workbook():
        path = fileopenbox(msg='Please select a ".xlsx" or "csv" file', title='Load File')
        try:
            wb = load_workbook(path)
        except:
            print('Selected file not ".xlsx" file type')
            exit()
        else:
            return wb

@unidentified_error_handler
def load_point_codes_list():
    prompt_continue('''    Select your code file.
    
    This file contains the point codes to be compared against (in column A)''')
    wb = select_workbook()
    ws = wb.active

    def valid_code(code):
        pattern = r'[0-9]'
        if re.search(pattern, code):
            return False
        else:
            return True
        
    codes = []
    column_to_check = 'A'
    invalid_codes_present = False
    invalid_codes = ''''''

    for code in ws[column_to_check]:
        if(code.value == None):
            continue
        if (valid_code(code.value) == False):
            invalid_codes = invalid_codes + f'WARNING Invalid Code: {code} {code.value}\n'
            invalid_codes_present = True
        codes.append(code.value)

    if (invalid_codes_present == True):
        prompt_continue(f"""Invalid codes present, continue?
{invalid_codes}""")
    return codes

@unidentified_error_handler
def load_survey_points():
    prompt_continue('''    Please select your survey file
                    
    This file can be raw from the surveyor, but must be in .xlsx format''')
    wb = select_workbook()
    ws = wb.active

    #parse survey points into lists [(pt_num, pt_desc)...] 
    def parse_points():
        colA = ws['A']
        colE = ws['E']
        error_list = []
        point_list = {}
        parsed_desc_point_list = []
        for num in range(len(colA)):
            pt_num = colA[num].value
            desc = colE[num].value

            def parse_description(desc):
                remove_numbers_and_decimals_pattern = r'-?\d+(\.\d+)?'
                remove_multiplication_symbol = r'\s[xX]\s'
                combined_pattern = r'|'.join((remove_multiplication_symbol, remove_numbers_and_decimals_pattern))
                sanitized_desc = re.sub(combined_pattern, '', desc)
                parsed_desc = sanitized_desc.split()
                return parsed_desc
            
            if (desc != None):
                try:
                    parsed_desc = parse_description(desc)
                    parsed_desc_point_list.append((pt_num, parsed_desc))
                    point_list[f'{pt_num}'] = desc
                except:
                    error_list.append((pt_num, desc))
            else:
                error_list.append((pt_num, None))

        return {'error_list': error_list, 'point_list': point_list, 'parsed_desc_point_list': parsed_desc_point_list}
    
    return parse_points()
    
@unidentified_error_handler
def check_descriptions_against_codes(codes, points):
    unknown_points_list = []
    for point in points['parsed_desc_point_list']:
        for desc in point[1]:
            if desc not in codes:
                unknown_points_list.append(point[0])
                continue
    return unknown_points_list

@unidentified_error_handler
def output_points(points, unknown_points):
    name = filesavebox(msg='Save file as .xlsx', title='Save Output To File', default='output')

    wb = Workbook()
    ws = wb.active

    for index, pt_num in enumerate(unknown_points, 1):
        point_list = points['point_list']
        pt_desc = point_list[f'{pt_num}']
        ws[f'A{index}'] = pt_num
        ws[f'E{index}'] = pt_desc
    
    wb.save(f'{name}.xlsx')
    
point_codes = load_point_codes_list()
survey_points = load_survey_points()
unknown_points = check_descriptions_against_codes(point_codes, survey_points)

output_points(survey_points, unknown_points)